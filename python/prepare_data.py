"""
prepare_data.py
===============
Convert the raw source datasets into clean, UTF-8 CSV files that Snowflake can
load with COPY INTO.

Why this step exists
--------------------
* The BLS OEWS files ship as Excel workbooks (.xlsx). Snowflake's COPY INTO
  cannot read .xlsx directly, so we flatten the single data sheet of each
  workbook into a comma-separated CSV.
* The O*NET files ("Task Ratings.txt" and "Task Statements.txt") are already
  plain text but are TAB separated. We copy them into data/raw unchanged so the
  loader has every input in one predictable place. (Snowflake handles the tab
  delimiter via the file format defined in the SQL scripts.)

This script NEVER touches Snowflake. It only reads the original downloads and
writes prepared copies into data/raw/. Run it once before loading.

Usage
-----
    python python/prepare_data.py                 # auto-detect source folder
    python python/prepare_data.py --source /path/to/data_warehouse
    python python/prepare_data.py --keep-zip-extract   # keep temp xlsx extracts

Outputs (written to ai_labor_snowflake/data/raw/):
    oews_2022.csv, oews_2023.csv, oews_2024.csv, oews_2025.csv
    onet_task_ratings.csv
    onet_task_statements.csv
    onet_occupation_data.csv   (SOC code, title, description - occupation context)
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import sys
import tempfile
import zipfile
from pathlib import Path

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
# This file lives at ai_labor_snowflake/python/prepare_data.py
PROJECT_ROOT = Path(__file__).resolve().parents[1]          # ai_labor_snowflake/
RAW_OUT_DIR = PROJECT_ROOT / "data" / "raw"

# The original downloads sit in the sibling "data_warehouse" folder by default.
# (…/Claude_Projects/data_warehouse). Adjust with --source if yours differ.
DEFAULT_SOURCE = PROJECT_ROOT.parent / "data_warehouse"

# Map each OEWS year to the zip that contains its workbook.
OEWS_ZIPS = {
    2022: "oesm22all.zip",
    2023: "oesm23all.zip",
    2024: "oesm24all.zip",
    2025: "oesm25all.zip",
}

# O*NET text bundle (contains Task Statements + Task Ratings, tab separated).
ONET_ZIP = "db_30_3_text.zip"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def find_source_dir(cli_source: str | None) -> Path:
    """Return the folder that holds the raw downloads."""
    if cli_source:
        p = Path(cli_source).expanduser().resolve()
        if not p.is_dir():
            sys.exit(f"ERROR: --source path does not exist: {p}")
        return p
    if DEFAULT_SOURCE.is_dir():
        return DEFAULT_SOURCE
    sys.exit(
        "ERROR: could not locate the raw data folder.\n"
        f"Looked for: {DEFAULT_SOURCE}\n"
        "Pass the folder explicitly with:  python python/prepare_data.py --source <folder>"
    )


def require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        sys.exit(
            "ERROR: openpyxl is required to read the OEWS .xlsx files.\n"
            "Install it with:  pip install openpyxl"
        )


def convert_oews_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    """
    Flatten the OEWS data sheet into a CSV.

    The workbook has several sheets; the data is always on the sheet whose name
    starts with "All " (e.g. "All May 2025 data"). We fall back to the first
    sheet if that pattern is not found.

    Returns the number of DATA rows written (excludes the header row).
    """
    import openpyxl

    # read_only + data_only keeps memory low and resolves formula values.
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if s.lower().startswith("all ")),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    data_rows = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        # QUOTE_MINIMAL + a real CSV writer safely handles the commas that live
        # inside OCC_TITLE / AREA_TITLE / NAICS_TITLE values.
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Normalise every cell to a string; treat None as empty so the
            # numeric-cleaning logic in the staging SQL sees a NULL-able blank.
            out = ["" if v is None else str(v) for v in row]
            writer.writerow(out)
            if i > 0:
                data_rows += 1
    wb.close()
    return data_rows


def copy_onet_txt(src_txt: Path, dest_csv: Path) -> int:
    """
    Copy an O*NET tab-separated file into data/raw unchanged (still tab
    separated). We keep the .csv extension for consistency but the SQL file
    format declares FIELD_DELIMITER = '\\t'.

    Returns number of data rows (excludes header).
    """
    shutil.copyfile(src_txt, dest_csv)
    # Count rows for the run summary.
    with open(dest_csv, "r", encoding="utf-8", errors="replace") as fh:
        total = sum(1 for _ in fh)
    return max(total - 1, 0)


def extract_member(zip_path: Path, name_contains: str, dest_dir: Path) -> Path:
    """Extract the first member of a zip whose name contains name_contains."""
    with zipfile.ZipFile(zip_path) as zf:
        members = [m for m in zf.namelist() if name_contains.lower() in m.lower()]
        if not members:
            raise FileNotFoundError(
                f"No member matching '{name_contains}' inside {zip_path.name}"
            )
        member = members[0]
        zf.extract(member, dest_dir)
        return dest_dir / member


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--source", help="Folder containing the raw downloads "
                                          "(zips + Task Ratings.txt).")
    parser.add_argument("--keep-zip-extract", action="store_true",
                        help="Keep the temporary extracted xlsx/txt files.")
    args = parser.parse_args()

    source = find_source_dir(args.source)
    RAW_OUT_DIR.mkdir(parents=True, exist_ok=True)
    require_openpyxl()

    print(f"Source folder : {source}")
    print(f"Output folder : {RAW_OUT_DIR}")
    print("-" * 60)

    tmp_dir = Path(tempfile.mkdtemp(prefix="ai_labor_prep_"))
    summary: list[tuple[str, int]] = []

    try:
        # ---- 1. OEWS (BLS) workbooks -> CSV -----------------------------
        for year, zip_name in OEWS_ZIPS.items():
            zip_path = source / zip_name
            if not zip_path.exists():
                print(f"[skip] {zip_name} not found - skipping OEWS {year}")
                continue
            print(f"[oews {year}] extracting {zip_name} ...")
            xlsx_path = extract_member(zip_path, ".xlsx", tmp_dir)
            out_csv = RAW_OUT_DIR / f"oews_{year}.csv"
            print(f"[oews {year}] converting {xlsx_path.name} -> {out_csv.name} ...")
            rows = convert_oews_xlsx_to_csv(xlsx_path, out_csv)
            print(f"[oews {year}] wrote {rows:,} data rows")
            summary.append((out_csv.name, rows))

        # ---- 2. O*NET Task Ratings --------------------------------------
        # Prefer the standalone "Task Ratings.txt" if present (it is the newest
        # / most complete extract). Fall back to the copy inside the O*NET zip.
        ratings_src = source / "Task Ratings.txt"
        if not ratings_src.exists():
            onet_zip = source / ONET_ZIP
            if onet_zip.exists():
                print(f"[ratings] extracting Task Ratings from {ONET_ZIP} ...")
                ratings_src = extract_member(onet_zip, "Task Ratings.txt", tmp_dir)
        if ratings_src and Path(ratings_src).exists():
            out = RAW_OUT_DIR / "onet_task_ratings.csv"
            print(f"[ratings] copying {Path(ratings_src).name} -> {out.name} ...")
            rows = copy_onet_txt(Path(ratings_src), out)
            print(f"[ratings] wrote {rows:,} data rows")
            summary.append((out.name, rows))
        else:
            print("[skip] Task Ratings not found")

        # ---- 3. O*NET Task Statements (inside db_30_3_text.zip) ----------
        onet_zip = source / ONET_ZIP
        if onet_zip.exists():
            print(f"[statements] extracting Task Statements from {ONET_ZIP} ...")
            stmt_src = extract_member(onet_zip, "Task Statements.txt", tmp_dir)
            out = RAW_OUT_DIR / "onet_task_statements.csv"
            print(f"[statements] copying {stmt_src.name} -> {out.name} ...")
            rows = copy_onet_txt(stmt_src, out)
            print(f"[statements] wrote {rows:,} data rows")
            summary.append((out.name, rows))
        else:
            print(f"[skip] {ONET_ZIP} not found - cannot extract Task Statements")

        # ---- 4. O*NET Occupation Data (inside db_30_3_text.zip) ----------
        # SOC code + occupation title + full description. Needed so the AI-score
        # prompt has occupational context (the same task means different work in
        # different occupations). Tab separated, 3 columns.
        if onet_zip.exists():
            print(f"[occdata] extracting Occupation Data from {ONET_ZIP} ...")
            occ_src = extract_member(onet_zip, "Occupation Data.txt", tmp_dir)
            out = RAW_OUT_DIR / "onet_occupation_data.csv"
            print(f"[occdata] copying {occ_src.name} -> {out.name} ...")
            rows = copy_onet_txt(occ_src, out)
            print(f"[occdata] wrote {rows:,} data rows")
            summary.append((out.name, rows))
        else:
            print(f"[skip] {ONET_ZIP} not found - cannot extract Occupation Data")

    finally:
        if not args.keep_zip_extract:
            shutil.rmtree(tmp_dir, ignore_errors=True)
        else:
            print(f"\nTemporary extracts kept in: {tmp_dir}")

    # ---- Summary --------------------------------------------------------
    print("\n" + "=" * 60)
    print("PREPARATION COMPLETE")
    print("=" * 60)
    if not summary:
        print("No files were produced. Check --source path and downloads.")
        sys.exit(1)
    for name, rows in summary:
        print(f"  {name:<28} {rows:>12,} rows")
    print(f"\nAll files ready in: {RAW_OUT_DIR}")
    print("Next: run python/load_local_files.py (after running sql/00 and sql/01).")


if __name__ == "__main__":
    main()
